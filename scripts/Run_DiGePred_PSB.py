#!/usr/bin/env python

import os
import json
import pandas as pd
import numpy as np
from scipy import stats
import _pickle as pickle
import itertools
import argparse
import datetime
from scipy import stats
import seaborn
import matplotlib as mpl
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import math
import sys
import codecs
import plotly.io as pio
import plotly.graph_objects as go
import re

sys.stdout = codecs.getwriter("ASCII")(sys.stdout.detach())

now = datetime.datetime.now()
month = str(now.strftime("%m"))
day = str(now.strftime("%d"))
year = str(now.strftime("%Y"))
hour = str(now.strftime("%H"))
minute = str(now.strftime("%M"))

## define output folders

out_root_folder = '/dors/capra_lab/projects/psb_collab/UDN/'
dependencies_location = '/dors/capra_lab/projects/psb_collab/UDN/files_for_digenic_prediction'
processed_sys_bio_data_location = '/dors/capra_lab/projects/psb_collab/UDN/data_for_systems_biology_analysis'
current_folder = os.getcwd()
out_dir = ''

## load ref dicts

refseq_to_gene_dict = pickle.load(open(os.path.join(dependencies_location, 'refseq_to_hgnc.pkl'), 'rb'))

alias_to_approved_name_dict = pickle.load(open(os.path.join(dependencies_location,
                                                            'alias_to_approved_gene_symbol_dict.pkl'), 'rb'))
approved_dict_to_alias_dict = pickle.load(open(os.path.join(dependencies_location,
                                                            'approved_symbol_to_alias_dict.pkl'), 'rb'))

pair_to_int_code_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'Network_Analysis', 'info_dicts', 'interaction_to_int_code_dict.pkl'), 'rb'))
interaction_details_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'Network_Analysis', 'info_dicts', 'interaction_code_details.pkl'), 'rb'))
int_code_to_pub_code_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'Network_Analysis', 'info_dicts', 'int_code_to_pub_id.pkl'), 'rb'))
with open(os.path.join(processed_sys_bio_data_location, 'Network_Analysis', 'info_dicts', 'pub_id_to_publication_details.pkl'), 'rb') as f:
    pub_id_details_dict = pickle.load(f)

kegg_gene_to_codes_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'Pathway_Enrichment', 'kegg_gene_to_path_codes.txt'), 'rb'))
kegg_code_to_name_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'Pathway_Enrichment', 'kegg_path_code_to_name.txt'), 'rb'))

reactome_gene_to_codes_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'Pathway_Enrichment', 'reactome_gene_to_path_codes.txt'), 'rb'))
reactome_code_to_name_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'Pathway_Enrichment', 'reactome_path_code_to_name.txt'), 'rb'))

pheno_gene_to_codes_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'HPO', 'hpo_gene_to_code.txt'), 'rb'))
pheno_code_to_name_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'HPO', 'updated_HPO_code_to_name_dict.pkl'), 'rb'))

pheno_tree = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'HPO', 'HPO_network_directed_all_shortest_path_lengths.pkl'), 'rb'))

gene_to_mean_TPM_by_tissue_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'GTEx', 'gene_to_mean_TPM_by_tissue_dict.pkl'), 'rb'))
id_to_broad_tissue_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'GTEx', 'id_to_broad_tissue_dict.pkl'), 'rb'))
id_to_narrow_tissue_dict = pickle.load(open(
    os.path.join(processed_sys_bio_data_location, 'GTEx', 'id_to_narrow_tissue_dict.pkl'), 'rb'))

# csvs_path = os.path.join(dependencies_location, 'all_human_gene_pairs_digenic_preds_csvs')
csvs_path = os.path.join(dependencies_location, 'all_human_gene_pairs_replace_ind_gene_feats_quad_mean')

digenic_csv_df = pd.read_csv(os.path.join(dependencies_location, 'held_out_test_digenic_feat_pred.csv'))
digenic_csv_df = digenic_csv_df.append(pd.read_csv(os.path.join(dependencies_location, 'train_digenic_feat_pred.csv')),
                                       ignore_index=True)

## read CSVs with DiGePred scores

csv_files = []
for r, d, f in os.walk(csvs_path):
    for file in sorted(f):
        csv_files.append(os.path.join(r, file))

pred_value_percentile_dict = pickle.load(open(os.path.join(dependencies_location,
                                                           'all_human_gene_pairs_pred_value_percentile_dict.pkl'),
                                              'rb'))
percentile_pred_value_dict = pickle.load(open(os.path.join(dependencies_location,
                                                           'all_human_gene_pairs_percentile_pred_value_dict.pkl'),
                                              'rb'))

pair_to_file_num_dict = pickle.load(open(os.path.join(dependencies_location, 'pair_to_file_num_dict.pkl'), 'rb'))
file_num_to_pair_dict = pickle.load(open(os.path.join(dependencies_location, 'file_to_num_to_pair_dict.pkl'), 'rb'))

three_aa_to_one_aa = {'ALA': 'A',
                      'ARG': 'R',
                      'ASN': 'N',
                      'ASP': 'D',
                      'CYS': 'C',
                      'GLN': 'Q',
                      'GLU': 'E',
                      'GLY': 'G',
                      'HIS': 'H',
                      'ILE': 'I',
                      'LEU': 'L',
                      'LYS': 'K',
                      'MET': 'M',
                      'PHE': 'F',
                      'PRO': 'P',
                      'SER': 'S',
                      'THR': 'T',
                      'TRP': 'W',
                      'TYR': 'Y',
                      'VAL': 'V',
                      }

parser = argparse.ArgumentParser(description='Get digenic score and percentile for gene pair')

parser.add_argument('-g', '--genes', type=str,
                    help='genes to get digenic preds',
                    dest='genes', required=False,
                    metavar='')

parser.add_argument('-e', '--excel', type=str,
                    help='UDN excel file to get digenic preds',
                    dest='excel', required=False,
                    metavar='')

parser.add_argument('-c', '--csv', type=str,
                    help='gene, tscript, cDNA variant, protein variant, zyg, inh \nin csv or tsv format to get digenic preds',
                    dest='csv', required=False,
                    metavar='')

parser.add_argument('-p', '--pairs', type=str,
                    help='pairs to get digenic preds',
                    dest='pairs', required=False,
                    metavar='')

parser.add_argument('-n', '--name', type=str,
                    help='analysis name or ID',
                    dest='name', required=False,
                    metavar='')

parser.add_argument('-d', '--outdir', type=str,
                    help='user output dir; default to current directory',
                    dest='dir', required=False,
                    metavar='')

parser.add_argument('-u', '--uniquekey', type=str,
                    help='unique key',
                    dest='uniquekey', required=False,
                    metavar='')

## parse user input

out_dir = ''
if parser.parse_args().dir is not None:
    usr_out_dir = parser.parse_args().dir

    print('User specified output directory: {}'.format(str(usr_out_dir)))

    if not os.path.isdir(usr_out_dir):
        os.mkdir(usr_out_dir)
        print('Folder created: {}'.format(usr_out_dir))
    else:
        print('Folder exists: {}'.format(usr_out_dir))

    out_dir = usr_out_dir
else:
    out_dir = current_folder
    print('output directory not specified; current directory is output directory:\n{n}'.format(n=current_folder))

if parser.parse_args().name is not None:
    project_name = parser.parse_args().name
    print('User specified project name: {n}'.format(n=project_name))

elif parser.parse_args().excel is not None:
    excel_file_name = parser.parse_args().excel

    wb = load_workbook(filename=excel_file_name)
    ws = wb.active
    parsed_excel_file = []
    for row in ws.iter_rows():
        parsed_excel_file.append([cell.value for cell in row])

    n = parsed_excel_file[0][0].find('UDN')
    case_id = str(parsed_excel_file[0][0][n:n + 9])
    project_name = case_id
    print('No name specified for project; UDN id derived from UDN excel file:\n{n}'.format(n=project_name))

else:
    date_id = '{y}-{m}-{d}-{hr}{mi}'.format(m=month, d=day, y=year, hr=hour, mi=minute)
    out_dir = os.path.join(out_root_folder, date_id)
    project_name = date_id
    print('No project name or ID found; Current local time used as project name:\n{n}'.format(n=project_name))

full_csv_save_path = os.path.join(out_dir, '{n}_all_gene_pairs_digenic_metrics.csv'
                                  .format(n=project_name))
sys_bio_details_graph_name = os.path.join(out_dir, '{n}_all_gene_pairs_sys_bio_details'
                                          .format(n=project_name))
digenic_score_graph_name = os.path.join(out_dir, '{n}_all_gene_pairs_digenic_score'
                                        .format(n=project_name))

html_file_name = os.path.join(out_dir, '{n}_all_gene_pairs_summary.html'
                              .format(n=project_name))

details_json_name = os.path.join(out_dir, '{n}_all_gene_pairs_summary.json'
                                 .format(n=project_name))

print(out_dir)

genes = []
pairs = []

gene_var_details_dict = dict()
gene_disp_name_to_ref_name_dict = dict()
gene_ref_name_to_disp_name_dict = dict()

isoform_counter_dict = dict()

if parser.parse_args().excel:

    excel_file_name = parser.parse_args().excel

    wb = load_workbook(filename=excel_file_name)
    ws = wb.active
    parsed_excel_file = []
    for row in ws.iter_rows():
        parsed_excel_file.append([cell.value for cell in row])

    n = parsed_excel_file[0][0].find('UDN')
    case_id = str(parsed_excel_file[0][0][n:n + 9])

    mut_marker = '\u25cf'

    genes = []
    label_row = []
    rels = []
    mom_col = 0
    dad_col = 0
    isoform_count = dict()
    for r_num, row in enumerate(parsed_excel_file):
        if row[0] is not None:
            if 'gene' in str(row[0]).lower():
                label_row = row
                for c_num, col in enumerate(label_row):
                    if col is not None:
                        if 'proband' in str(col).lower():
                            proband_col = c_num
                        if 'mother' in str(col).lower() or 'mom' in str(col).lower():
                            mom_col = c_num
                            rels.append('mom')
                        if 'father' in str(col).lower() or 'dad' in str(col).lower():
                            dad_col = c_num
                            rels.append('dad')
                break

    for r_num, row in enumerate(parsed_excel_file):
        if row[0] is not None:
            inh = []
            z = ''
            if row[proband_col] is not None:
                if '\u25cf' in row[proband_col] or '\u006f' in row[proband_col] or '-' in row[proband_col] or '\u25cb' in row[proband_col]:
                    if mom_col != 0 and row[mom_col] is not None:
                        if '\u25cf' in row[mom_col]:
                            inh.append('mom')
                        elif str(row[proband_col]) == str(row[mom_col]):
                            inh.append('mom')

                    if dad_col != 0 and row[dad_col] is not None:
                        if '\u25cf' in row[dad_col]:
                            inh.append('dad')
                        elif str(row[proband_col]) == str(row[dad_col]):
                            inh.append('dad')

                    if str(row[proband_col]).lower().count('\u25cf') == 1:
                        z = 'het'
                    elif str(row[proband_col]).lower().count('\u25cf') >= 2:
                        z = 'homo'

                    if 'NM_' in row[0]:
                        t = str(row[0])[str(row[0]).find('NM_'):].strip().rstrip().replace('_x000D_', '').split()[
                            0].strip().rstrip()
                        g = row[0].split('NM_')[0].strip().rstrip().replace('_x000D_', '').split()[0].strip().rstrip()

                    elif 'NM ' in row[0]:
                        r = str(row[0]).replace('NM ', 'NM_')
                        t = r[r.find('NM_'):].strip().rstrip().replace('_x000D_', '').split()[0].strip().rstrip()
                        g = r.split('NM_')[0].strip().rstrip().replace('_x000D_', '').split()[0].strip().rstrip()

                    else:
                        t = 'NA'
                        g = str(row[0]).split()[0].replace('_x000D_', '').strip().rstrip().split()[0].strip().rstrip()

                    if ':' not in g and g.lower() != 'gene':

                        chr = parsed_excel_file[r_num][1]
                        pos = parsed_excel_file[r_num + 1][1]
                        rsid = [parsed_excel_file[r_num + 1][2] if parsed_excel_file[r_num + 1][2] is not None
                                else ''][0]
                        ref = [parsed_excel_file[r_num][2].split('→')[0].strip().rstrip() if
                               len(parsed_excel_file[r_num]) > 2 and parsed_excel_file[r_num][2] is not None and
                               '→' in parsed_excel_file[r_num][2] else ''][0]
                        alt = [parsed_excel_file[r_num][2].split('→')[1].strip().rstrip() if
                               len(parsed_excel_file[r_num]) > 2 and parsed_excel_file[r_num][2] is not None and
                               '→' in parsed_excel_file[r_num][2] else ''][0]
                        consequence = [parsed_excel_file[r_num][3] if len(parsed_excel_file[r_num]) > 3 else ''][0]

                        if parsed_excel_file[r_num + 2][2] is not None and 'p.' in parsed_excel_file[r_num + 2][2]:
                            mut_long = parsed_excel_file[r_num + 2][2]
                            new_string = mut_long
                            for i in range(len(mut_long) - 2):
                                code = mut_long[i:i + 3]
                                if code.upper() in three_aa_to_one_aa:
                                    one_letter_code = three_aa_to_one_aa[code.upper()]
                                    new_string = new_string.replace(code, one_letter_code)
                            pv = new_string
                        else:
                            pv = ''

                        if parsed_excel_file[r_num + 1][2] is not None and 'c.' in parsed_excel_file[r_num + 1][2]:
                            cv = parsed_excel_file[r_num + 1][2]
                        else:
                            cv = ''

                        if not inh:
                            if len(set(rels).intersection(['mom', 'dad'])) < 2:
                                inh = 'both parents not sequenced'
                            else:
                                inh = ['de novo']
                        elif 'mom' in inh and 'dad' in inh:
                            inh = ['mom, dad']

                        derived_g_names = [refseq_to_gene_dict[t] if t != '' and t in refseq_to_gene_dict else ['']][0]

                        if '' in derived_g_names or g in derived_g_names:
                            disp_name = g
                            if disp_name not in genes:
                                if disp_name not in alias_to_approved_name_dict and disp_name in approved_dict_to_alias_dict:
                                    ref_name = g
                                elif disp_name in alias_to_approved_name_dict and disp_name not in approved_dict_to_alias_dict:
                                    ref_name = alias_to_approved_name_dict[disp_name]
                                genes.append(disp_name)
                                isoform_counter_dict[disp_name] = 2
                                gene_disp_name_to_ref_name_dict[disp_name] = ref_name
                                gene_ref_name_to_disp_name_dict[ref_name] = disp_name
                            elif disp_name in genes and t != 'NA' and t not in gene_var_details_dict[disp_name]['tscript']:
                                disp_name2 = disp_name + '_{}'.format(isoform_counter_dict[g])
                                gene_disp_name_to_ref_name_dict[disp_name2] = ref_name
                                gene_ref_name_to_disp_name_dict[ref_name] = disp_name2
                                isoform_counter_dict[disp_name] += 1
                                ref_name = gene_disp_name_to_ref_name_dict[disp_name]
                                genes.append(disp_name2)

                        elif g not in derived_g_names:
                            g2 = derived_g_names[0]
                            disp_name = g2 + '({})'.format(g)
                            if g2 not in alias_to_approved_name_dict and g2 in approved_dict_to_alias_dict:
                                ref_name = disp_name
                            elif g2 in alias_to_approved_name_dict and g2 not in approved_dict_to_alias_dict:
                                ref_name = alias_to_approved_name_dict[g2]
                            isoform_counter_dict[disp_name] = 2
                            genes.append(disp_name)
                            gene_disp_name_to_ref_name_dict[disp_name] = ref_name
                            gene_ref_name_to_disp_name_dict[ref_name] = disp_name

                        if disp_name not in gene_var_details_dict:
                            gene_var_details_dict[disp_name] = {'tscript': t,
                                                                'chr': '',
                                                                'pos': [],
                                                                'ref': [],
                                                                'alt': [],
                                                                'rs ID': [],
                                                                'consequence': [],
                                                                'cdna var': [],
                                                                'prot var': [],
                                                                'zyg': [],
                                                                'inh': []
                                                                }

                        gene_var_details_dict[disp_name]['chr'] = chr
                        gene_var_details_dict[disp_name]['pos'].append(pos)
                        gene_var_details_dict[disp_name]['ref'].append(ref)
                        gene_var_details_dict[disp_name]['alt'].append(alt)
                        gene_var_details_dict[disp_name]['rs ID'].append(rsid)
                        gene_var_details_dict[disp_name]['consequence'].append(consequence)
                        gene_var_details_dict[disp_name]['cdna var'].append(cv)
                        gene_var_details_dict[disp_name]['prot var'].append(pv)
                        gene_var_details_dict[disp_name]['zyg'].append(z)
                        gene_var_details_dict[disp_name]['inh'].append(inh)

    pairs = itertools.combinations(set(genes), 2)

elif parser.parse_args().csv:
    print('Found gene transcript variant zygosity inheritance csv / tsv file')
    gene_var_inh_file = open(parser.parse_args().csv).read().split('\n')[:-1]

    isoform_count = dict()

    for line in gene_var_inh_file:
        if ',' in line:
            x = line.split(',')
        elif '\t' in line:
            x = line.split('\t')
        else:
            print('gene variant inheritance file not csv / tsv')
            break

        g = x[0].strip().rstrip()
        t = x[1].strip().rstrip()
        cv = x[2].strip().rstrip()
        pv = x[3].strip().rstrip()
        z = x[4].strip().rstrip()
        inh = x[5].strip().rstrip()

        genes.append(g)

        if g not in gene_var_details_dict:
            gene_var_details_dict[g] = {'tscript': t,
                                        'cdna var': [],
                                        'prot var': [],
                                        'zyg': [],
                                        'inh': []
                                         }

            gene_var_details_dict[g] = {'tscript': t,
                                        'cdna var': [],
                                        'prot var': [],
                                        'zyg': [],
                                        'inh': []
                                        }

            gene_var_details_dict[g]['cdna var'].append(cv)
            gene_var_details_dict[g]['prot var'].append(pv)
            gene_var_details_dict[g]['zyg'].append(z)
            gene_var_details_dict[g]['inh'].append(inh)

    pairs = itertools.combinations(set(genes), 2)

elif parser.parse_args().pairs:
    print('Found user input gene pairs file')
    pairs_list = open(parser.parse_args().pairs).read().split('\n')[:-1]

    for line in pairs_list:
        g1 = line.split(',')[0].strip().rstrip()
        g2 = line.split(',')[1].strip().rstrip()

        pairs.append(tuple(sorted([g1, g2])))

    pairs = sorted(set(pairs))

elif parser.parse_args().genes:
    print('Found user input genes file')

    genes = sorted(set(open(parser.parse_args().genes).read().split('\n')[:-1]))
    pairs = itertools.combinations(set(genes), 2)

else:
    print('No input file found!')


def get_digenic_metrics_csv(gene_pairs): ## generate DiGePred scores and gene pair details DataFrame
    gene_pairs = set([tuple(sorted(gp)) for gp in gene_pairs])
    print(len(gene_pairs))

    genes = []
    for gp in gene_pairs:
        genes.extend(list(gp))

    fnum_sel_pairs_dict = dict()
    fnum_sel_geneA_dict = dict()
    fnum_sel_geneB_dict = dict()
    disp_gp_dict = dict()

    for p in sorted(gene_pairs):
        gp = tuple(sorted([p[0], p[1]]))
        cgp = tuple(sorted([gene_disp_name_to_ref_name_dict[p[0]], gene_disp_name_to_ref_name_dict[p[1]]]))

        for fp, lp in pair_to_file_num_dict:
            if fp <= cgp <= lp:

                file_num = pair_to_file_num_dict[(fp, lp)]
                if file_num not in fnum_sel_pairs_dict:
                    fnum_sel_pairs_dict[file_num] = []
                    disp_gp_dict[file_num] = []
                    fnum_sel_geneA_dict[file_num] = []
                    fnum_sel_geneB_dict[file_num] = []
                fnum_sel_pairs_dict[file_num].append(cgp)
                disp_gp_dict[file_num].append(gp)
                fnum_sel_geneA_dict[file_num].append(cgp[0])
                fnum_sel_geneB_dict[file_num].append(cgp[1])

    sel_gp_csv_df = pd.DataFrame()

    for fnum in fnum_sel_pairs_dict:
        pairs = fnum_sel_pairs_dict[fnum]
        # print('File # {n}:'.format(n=fnum))

        num_csv_df = pd.read_csv([x for x in csv_files if '_{}.csv'.format(fnum) in x][0])
        gp_csv_df = num_csv_df[num_csv_df['gene A'].isin(fnum_sel_geneA_dict[fnum]) &
                               num_csv_df['gene B'].isin(fnum_sel_geneB_dict[fnum])]

        sel_gp_csv_df = sel_gp_csv_df.append(gp_csv_df, ignore_index=True)

    percentiles = []
    pwy_sims = []
    pheno_sims = []
    ppi_dists = []
    pwy_dists = []
    lit_dists = []
    coex_ranks = []
    genes_A = []
    genes_B = []
    tscripts_A = []
    tscripts_B = []
    zygs_A = []
    zygs_B = []
    cdna_vars_A = []
    cdna_vars_B = []
    prot_vars_A = []
    prot_vars_B = []
    inhs_A = []
    inhs_B = []
    gp_inhs = []
    preds = []

    for i in sorted(sel_gp_csv_df.index):

        ga = gene_ref_name_to_disp_name_dict[sel_gp_csv_df.iloc[i]['gene A']]
        gb = gene_ref_name_to_disp_name_dict[sel_gp_csv_df.iloc[i]['gene B']]

        genes_A.append(ga)
        genes_B.append(gb)

        pred = '%0.3g' % float(sel_gp_csv_df.iloc[i]['all digenic vs unaffected score'])
        preds.append(pred)
        pwy_sims.append(sel_gp_csv_df.iloc[i]['pathway similarity'] * 100)
        pheno_sims.append(sel_gp_csv_df.iloc[i]['phenotype similarity'] * 100)
        ppi_dists.append([1. / sel_gp_csv_df.iloc[i]['PPI distance']
                          if sel_gp_csv_df.iloc[i]['PPI distance'] > 0
                          else 0][0])
        pwy_dists.append([1. / sel_gp_csv_df.iloc[i]['pathway distance']
                          if sel_gp_csv_df.iloc[i]['pathway distance'] > 0
                          else 0][0])
        lit_dists.append([1. / sel_gp_csv_df.iloc[i]['literature distance']
                          if sel_gp_csv_df.iloc[i]['literature distance'] > 0
                          else 0][0])
        coex_ranks.append([1. / sel_gp_csv_df.iloc[i]['mutual co-expression rank']
                           if sel_gp_csv_df.iloc[i]['mutual co-expression rank'] > 0
                           else 0][0])

        res = True
        for i in range(len(pred_value_percentile_dict) - 1):
            lower = float(list(pred_value_percentile_dict)[i])
            upper = float(list(pred_value_percentile_dict)[i + 1])
            if lower <= float(pred) < upper:
                ptile = pred_value_percentile_dict[list(pred_value_percentile_dict)[i]]
                res = False
                break

        if res:
            ptile = 100

        percentiles.append(ptile)

        if ga in gene_var_details_dict:
            t_A = gene_var_details_dict[ga]['tscript']
            ii_A = gene_var_details_dict[ga]['inh']
            cv_A = gene_var_details_dict[ga]['cdna var']
            pv_A = gene_var_details_dict[ga]['prot var']
            z_A = gene_var_details_dict[ga]['zyg']
        else:
            inhs_A.append(['NA'])
            tscripts_A.append(['NA'])
            cdna_vars_A.append(['NA'])
            prot_vars_A.append(['NA'])
            zygs_A.append(['NA'])

        if gb in gene_var_details_dict:
            t_B = gene_var_details_dict[gb]['tscript']
            ii_B = gene_var_details_dict[gb]['inh']
            cv_B = gene_var_details_dict[gb]['cdna var']
            pv_B = gene_var_details_dict[gb]['prot var']
            z_B = gene_var_details_dict[gb]['zyg']
        else:
            inhs_B.append(['NA'])
            tscripts_B.append(['NA'])
            cdna_vars_B.append(['NA'])
            prot_vars_B.append(['NA'])
            zygs_B.append(['NA'])

        if ga in gene_var_details_dict and gb in gene_var_details_dict:

            mom_inh_A = False
            dad_inh_A = False
            de_novo_A = False

            if 'mom' in str(ii_A) or 'mother' in str(ii_A) or 'maternal' in str(ii_A):
                mom_inh_A = True
            if 'dad' in str(ii_A) or 'father' in str(ii_A) or 'paternal' in str(ii_A):
                dad_inh_A = True
            if 'de novo' in str(ii_A):
                de_novo_A = True

            inh_A = 'NA'
            if mom_inh_A or dad_inh_A:
                if mom_inh_A and dad_inh_A:
                    inh_A = 'comp het'
                elif mom_inh_A:
                    inh_A = 'maternal'
                elif dad_inh_A:
                    inh_A = 'paternal'
            elif 'both parents not sequenced' not in str(ii_A):
                inh_A = 'de novo'
            else:
                inh_A = 'both parents not sequenced'

            if de_novo_A and 'de novo' not in inh_A:
                inh_A = inh_A + ' de novo'

            mom_inh_B = False
            dad_inh_B = False
            de_novo_B = False

            if 'mom' in str(ii_B) or 'mother' in str(ii_B) or 'maternal' in str(ii_B):
                mom_inh_B = True
            if 'dad' in str(ii_B) or 'father' in str(ii_B) or 'paternal' in str(ii_B):
                dad_inh_B = True
            if 'de novo' in str(ii_B):
                de_novo_B = True

            inh_B = 'NA'
            if mom_inh_B or dad_inh_B:
                if mom_inh_B and dad_inh_B:
                    inh_B = 'comp het'
                elif mom_inh_B:
                    inh_B = 'maternal'
                elif dad_inh_B:
                    inh_B = 'paternal'
            elif 'both parents not sequenced' not in str(ii_B):
                inh_B = 'de novo'
            else:
                inh_B = 'both parents not sequenced'

            if de_novo_B and 'de novo' not in inh_B:
                inh_B = inh_B + ' de novo'

            gp_inh = 'NA'

            if inh_A != 'NA' and inh_B != 'NA':
                if 'comp het' in str([inh_A] + [inh_B]):
                    if 'de novo' in str([inh_A] + [inh_B]):
                        gp_inh = 'bi-parental + de novo'
                    else:
                        gp_inh = 'bi-parental'

                elif 'maternal' in str([inh_A] + [inh_B]) and 'paternal' in str([inh_A] + [inh_B]):
                    if 'de novo' in str([inh_A] + [inh_B]):
                        gp_inh = 'bi-parental + de novo'
                    else:
                        gp_inh = 'bi-parental'

                elif 'de novo' in str([inh_A] + [inh_B]):
                    gp_inh = 'de novo'

                elif 'maternal' in inh_A and 'maternal' in inh_B:
                    gp_inh = 'maternal'
                elif 'paternal' in inh_A and 'paternal' in inh_B:
                    gp_inh = 'paternal'
                elif 'both parents not sequenced' in inh_A or 'both parents not sequenced' in inh_B:
                    gp_inh = 'both parents not sequenced'

                inhs_A.append((inh_A, ii_A))
                inhs_B.append((inh_B, ii_B))
                gp_inhs.append(gp_inh)

            else:
                inhs_A.append(['NA'])
                inhs_B.append(['NA'])
                gp_inhs.append(['NA'])

            tscripts_A.append(t_A)
            tscripts_B.append(t_B)

            if cv_A:
                cdna_vars_A.append(cv_A)
            else:
                cdna_vars_A.append(['NA'])
            if cv_B:
                cdna_vars_B.append(cv_B)
            else:
                cdna_vars_B.append(['NA'])

            if pv_A:
                prot_vars_A.append(pv_A)
            else:
                prot_vars_A.append(['NA'])
            if pv_B:
                prot_vars_B.append(pv_B)
            else:
                prot_vars_B.append(['NA'])

            if z_A:
                zygs_A.append(z_A)
            else:
                zygs_A.append(['NA'])
            if z_B:
                zygs_B.append(z_B)
            else:
                zygs_B.append(['NA'])
        else:
            gp_inhs.append(['NA'])

    dig_df = pd.DataFrame(index=range(1, len(sel_gp_csv_df.index) + 1))
    dig_df['gene A'] = genes_A
    dig_df['gene B'] = genes_B
    dig_df['digenic score'] = preds
    dig_df['pair inheritance'] = gp_inhs
    dig_df['percentile'] = percentiles
    dig_df['transcript A'] = tscripts_A
    dig_df['cDNA change A'] = cdna_vars_A
    dig_df['protein change A'] = prot_vars_A
    dig_df['zygosity A'] = zygs_A
    dig_df['inheritance A'] = inhs_A
    dig_df['transcript B'] = tscripts_B
    dig_df['cDNB change B'] = cdna_vars_B
    dig_df['protein change B'] = prot_vars_B
    dig_df['zygosity B'] = zygs_B
    dig_df['inheritance B'] = inhs_B
    dig_df['pathway similarity'] = pwy_sims
    dig_df['phenotype similarity'] = pheno_sims
    dig_df['PPI distance'] = ppi_dists
    dig_df['pathway distance'] = pwy_dists
    dig_df['literature distance'] = lit_dists
    dig_df['mutual co-expression rank'] = coex_ranks

    print('\nDigenic metrics computed for all gene pairs\n')
    print(dig_df)

    dig_df.to_csv(full_csv_save_path, sep=',', header=True, index=False)

    print('output csv filename: {}'.format(full_csv_save_path))

    return dig_df, sel_gp_csv_df


def plot_heatmaps(dig_df, sel_gp_csv_df): ## plot DiGePred score and systems biology data heatmaps

    cmaps = dict()
    cmaps['ppi dist'] = 'Purples_r'
    cmaps['pwy dist'] = 'Purples_r'
    cmaps['txt dist'] = 'Purples_r'
    cmaps['pwy similarity'] = 'Purples'
    cmaps['pheno similarity'] = 'Purples'
    cmaps['dig score'] = 'Purples'
    cmaps['dig percentile'] = 'Purples'
    cmaps['coex rank'] = 'Purples_r'

    vmins = dict()
    vmins['ppi dist'] = 1
    vmins['pwy dist'] = 1
    vmins['txt dist'] = 1
    vmins['pwy similarity'] = 0.1
    vmins['pheno similarity'] = 0.1
    vmins['dig score'] = 0.098
    vmins['dig percentile'] = 0
    vmins['coex rank'] = 1

    vmaxs = dict()
    vmaxs['ppi dist'] = 1.001
    vmaxs['pwy dist'] = 1.001
    vmaxs['txt dist'] = 1.001
    vmaxs['pwy similarity'] = 40
    vmaxs['pheno similarity'] = 40
    vmaxs['dig score'] = 0.55
    vmaxs['dig percentile'] = 100
    vmaxs['coex rank'] = 500

    ticks = dict()
    ticks['ppi dist'] = []
    ticks['pwy dist'] = []
    ticks['txt dist'] = []
    ticks['pwy similarity'] = [30, 60, 90]
    ticks['pheno similarity'] = [30, 60, 90]
    ticks['dig score'] = [0.05, 0.1, 0.25, 0.5]
    ticks['dig percentile'] = [30, 60, 90]
    ticks['coex rank'] = [300, 600, 900]

    ax_labels = dict()
    ax_labels['ppi dist'] = 'PPI network interactions'
    ax_labels['pwy dist'] = 'pathways network interactions'
    ax_labels['txt dist'] = 'literature-mined network interactions'
    ax_labels['pwy similarity'] = '% pathways similarity'
    ax_labels['pheno similarity'] = '% phenotype similarity'
    ax_labels['dig score'] = 'digenic score \n(high confidence threshold = 0.5)'
    ax_labels['dig percentile'] = 'percentile of digenic score'
    ax_labels['coex rank'] = 'mutual co-expression rank'

    annots = dict()
    annots['ppi dist'] = '.0f'
    annots['pwy dist'] = '.0f'
    annots['txt dist'] = '.0f'
    annots['pwy similarity'] = '.0f'
    annots['pheno similarity'] = '.0f'
    annots['dig score'] = '.1g'
    annots['dig percentile'] = '.0f'
    annots['coex rank'] = '.0f'

    maternal_color = '#fee0b6'
    paternal_color = '#d9f0d3'
    bi_parental_color = '#1b7837'
    de_novo_color = '#b35806'
    bi_parental_de_novo_color = 'k'
    background_color = 'white'
    seq_mis_color = '#969696'

    genes = sorted(set(list(dig_df['gene A'])).union(list(dig_df['gene B'])))

    dfs = dict()
    dfs['ppi dist'] = pd.DataFrame(index=genes, columns=genes,
                                   data=np.zeros(shape=(len(genes), len(genes))))

    dfs['pwy dist'] = pd.DataFrame(index=genes, columns=genes,
                                   data=np.zeros(shape=(len(genes), len(genes))))

    dfs['txt dist'] = pd.DataFrame(index=genes, columns=genes,
                                   data=np.zeros(shape=(len(genes), len(genes))))

    dfs['pwy similarity'] = pd.DataFrame(index=genes, columns=genes,
                                         data=np.zeros(shape=(len(genes), len(genes))))

    dfs['pheno similarity'] = pd.DataFrame(index=genes, columns=genes,
                                           data=np.zeros(shape=(len(genes), len(genes))))

    dfs['coex rank'] = pd.DataFrame(index=genes, columns=genes,
                                    data=np.zeros(shape=(len(genes), len(genes))))

    dfs['dig score'] = pd.DataFrame(index=genes, columns=genes,
                                    data=np.zeros(shape=(len(genes), len(genes))))

    dfs['dig percentile'] = pd.DataFrame(index=genes, columns=genes,
                                         data=np.zeros(shape=(len(genes), len(genes))))

    for i in sorted(sel_gp_csv_df.index):

        g1 = sel_gp_csv_df.iloc[i]['gene A']
        g2 = sel_gp_csv_df.iloc[i]['gene B']

        if g1 != g2:
            ga, gb = tuple(sorted([g1, g2]))

            dfs['ppi dist'][ga][gb] = dig_df[(dig_df['gene A'] == ga) & (dig_df['gene B'] == gb)]['PPI distance']
            dfs['ppi dist'][gb][ga] = dfs['ppi dist'][ga][gb]

            dfs['pwy dist'][ga][gb] = dig_df.loc[(dig_df['gene A'] == ga) & (dig_df['gene B'] == gb)][
                'pathway distance']
            dfs['pwy dist'][gb][ga] = dfs['pwy dist'][ga][gb]

            dfs['txt dist'][ga][gb] = dig_df.loc[(dig_df['gene A'] == ga) & (dig_df['gene B'] == gb)][
                'literature distance']
            dfs['txt dist'][gb][ga] = dfs['txt dist'][ga][gb]

            dfs['pwy similarity'][ga][gb] = dig_df.loc[(dig_df['gene A'] == ga) & (dig_df['gene B'] == gb)][
                'pathway similarity']
            dfs['pwy similarity'][gb][ga] = dfs['pwy similarity'][ga][gb]

            dfs['pheno similarity'][ga][gb] = dig_df.loc[(dig_df['gene A'] == ga) & (dig_df['gene B'] == gb)][
                'phenotype similarity']
            dfs['pheno similarity'][gb][ga] = dfs['pheno similarity'][ga][gb]

            dfs['coex rank'][ga][gb] = dig_df.loc[(dig_df['gene A'] == ga) & (dig_df['gene B'] == gb)][
                'mutual co-expression rank']
            dfs['coex rank'][gb][ga] = dfs['coex rank'][ga][gb]

            dfs['dig score'][ga][gb] = float(dig_df.loc[(dig_df['gene A'] == ga) & (dig_df['gene B'] == gb)][
                                                 'digenic score'])
            dfs['dig score'][gb][ga] = dfs['dig score'][ga][gb]

            dfs['dig percentile'][ga][gb] = math.floor(dig_df.loc[(dig_df['gene A'] == ga) & (dig_df['gene B'] == gb)][
                                                           'percentile'])
            dfs['dig percentile'][gb][ga] = dfs['dig percentile'][ga][gb]

    fig, ax = plt.subplots(2, 3, figsize=(30, 20))
    for i, feat in enumerate([d for d in dfs if 'dig' not in d]):

        df_clean = dfs[feat].copy()
        if 'similarity' not in feat and 'dig' not in feat:
            df_clean[df_clean > vmaxs[feat]] = np.nan
        df_clean[df_clean < vmins[feat]] = np.nan

        col = i % 3
        row = i // 3

        mask = np.zeros_like(df_clean)
        mask[np.triu_indices_from(mask)] = True
        ax[row][col].patch.set_facecolor(background_color)

        annot_kws = dict()
        annot_kws['ppi dist'] = {"size": 0, "style": 'italic'}
        annot_kws['pwy dist'] = {"size": 0, "style": 'italic'}
        annot_kws['txt dist'] = {"size": 0, "style": 'italic'}
        annot_kws['pwy similarity'] = {"size": 25 * len(df_clean) ** -0.25, "style": 'italic'}
        annot_kws['pheno similarity'] = {"size": 25 * len(df_clean) ** -0.25, "style": 'italic'}
        annot_kws['dig score'] = {"size": 25 * len(df_clean) ** -0.25, "style": 'italic'}
        annot_kws['dig percentile'] = {"size": 25 * len(df_clean) ** -0.25, "style": 'italic'}
        annot_kws['coex rank'] = {"size": 25 * len(df_clean) ** -0.25, "style": 'italic'}

        hmap = seaborn.heatmap(df_clean, ax=ax[row][col], vmin=vmins[feat], vmax=vmaxs[feat],
                               linewidths=4, linecolor=background_color, cbar=False,
                               square=True, cmap=cmaps[feat], mask=mask, annot=True, fmt=annots[feat],
                               annot_kws=annot_kws[feat],
                               cbar_kws={"shrink": .4, 'ticks': ticks[feat], 'spacing': 'proportional'})

        ax[row][col].set_xticks([i + 0.5 for i in range(len(df_clean.index) - 1)])
        ax[row][col].set_xticklabels(df_clean.index[:-1], fontsize=12, rotation=70, ha='center')
        ax[row][col].set_yticks([i + 1.5 for i in range(len(df_clean.index) - 1)])
        ax[row][col].set_yticklabels(df_clean.index[1:], fontsize=12, rotation=30)
        ax[row][col].set_xlabel(ax_labels[feat], fontsize=16)
        ax[row][col].set_ylim([len(df_clean.index), 1])

        for x, g1 in enumerate(df_clean.index[:-1]):
            for y, g2 in enumerate(df_clean.index[1:]):

                if g1 != g2 and mask[x][y] == True:

                    ga, gb = tuple(sorted([g1, g2]))
                    inh = str(dig_df.loc[(dig_df['gene A'] == ga) & (dig_df['gene B'] == gb)]['pair inheritance'])

                    if 'bi-parental' in inh and 'de novo' in inh:

                        ax[row][col].axhspan(ymin=y + 1 + 0.04,
                                             ymax=y + 2 - 0.04,
                                             xmin=(x + 0.04) / len(df_clean),
                                             xmax=(x + 1 - 0.04) / len(df_clean), linewidth=1,
                                             edgecolor=bi_parental_de_novo_color, facecolor=None, fill=False,
                                             alpha=0.95)

                    elif 'bi-parental' in inh:
                        ax[row][col].axhspan(ymin=y + 1 + 0.04,
                                             ymax=y + 2 - 0.04,
                                             xmin=(x + 0.04) / len(df_clean),
                                             xmax=(x + 1 - 0.04) / len(df_clean), linewidth=1,
                                             edgecolor=bi_parental_color, facecolor=None, fill=False,
                                             alpha=0.95)
                    elif 'de novo' in inh:
                        ax[row][col].axhspan(ymin=y + 1 + 0.04,
                                             ymax=y + 2 - 0.04,
                                             xmin=(x + 0.04) / len(df_clean),
                                             xmax=(x + 1 - 0.04) / len(df_clean), linewidth=1,
                                             edgecolor=de_novo_color, facecolor=None, fill=False,
                                             alpha=0.95)

                    elif 'maternal' in inh:
                        ax[row][col].axhspan(ymin=y + 1 + 0.04,
                                             ymax=y + 2 - 0.04,
                                             xmin=(x + 0.04) / len(df_clean),
                                             xmax=(x + 1 - 0.04) / len(df_clean), linewidth=1,
                                             edgecolor=maternal_color, facecolor=None, fill=False,
                                             alpha=0.95)

                    elif 'paternal' in inh:
                        ax[row][col].axhspan(ymin=y + 1 + 0.04,
                                             ymax=y + 2 - 0.04,
                                             xmin=(x + 0.04) / len(df_clean),
                                             xmax=(x + 1 - 0.04) / len(df_clean), linewidth=1,
                                             edgecolor=paternal_color, facecolor=None, fill=False,
                                             alpha=0.95)
                    else:
                        ax[row][col].axhspan(ymin=y + 1 + 0.02,
                                             ymax=y + 2 - 0.02,
                                             xmin=(x + 0.02) / len(df_clean),
                                             xmax=(x + 1 - 0.02) / len(df_clean), linewidth=1,
                                             edgecolor=seq_mis_color, facecolor=None, fill=False,
                                             alpha=0.95)

    handles = []
    labels = []

    m_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                  edgecolor=maternal_color, facecolor=None, fill=False, linewidth=3, alpha=1)
    p_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                  edgecolor=paternal_color, facecolor=None, fill=False, linewidth=3, alpha=1)

    bp_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                   edgecolor=bi_parental_color, facecolor=None, fill=False, linewidth=3, alpha=1)
    dn_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                   edgecolor=de_novo_color, facecolor=None, fill=False, linewidth=3, alpha=1)
    both_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                     edgecolor=bi_parental_de_novo_color, facecolor=None, fill=False, linewidth=3,
                                     alpha=1)
    miss_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                     edgecolor=seq_mis_color, facecolor=None, fill=False, linewidth=1,
                                     alpha=1)

    handles.append(m_box)
    labels.append('maternal inheritance')
    handles.append(p_box)
    labels.append('paternal inheritance')
    handles.append(bp_box)
    labels.append('biparental inheritance')
    handles.append(dn_box)
    labels.append('de novo variants')
    handles.append(both_box)
    labels.append('biparental inheritance & de novo')
    handles.append(miss_box)
    labels.append('sequencing data missing for one or both parents')

    fig.legend(handles, labels, loc='upper center', fontsize=20, handlelength=0.7, ncol=3,
               facecolor=background_color, edgecolor=background_color, frameon=False, framealpha=0)
    fig.patch.set_facecolor(background_color)

    plt.subplots_adjust(left=0.125, bottom=0.1, right=0.9, top=1, wspace=0.1, hspace=0.1)
    plt.savefig(sys_bio_details_graph_name + '.pdf', bbox_inches='tight', facecolor=background_color, edgecolor='none')
    plt.savefig(sys_bio_details_graph_name + '.svg', bbox_inches='tight', facecolor=background_color, edgecolor='none')

    print('features heatmap filename: {}'.format(sys_bio_details_graph_name))

    fig, ax = plt.subplots(1, 1, figsize=(10, 10))
    feat = 'dig score'

    df_clean = dfs[feat].copy()
    df_clean[df_clean < vmins[feat]] = np.nan

    mask = np.zeros_like(df_clean)
    mask[np.triu_indices_from(mask)] = True
    ax.patch.set_facecolor(background_color)

    annot_kws = dict()
    annot_kws['ppi dist'] = {"size": 0, "style": 'italic'}
    annot_kws['pwy dist'] = {"size": 0, "style": 'italic'}
    annot_kws['txt dist'] = {"size": 0, "style": 'italic'}
    annot_kws['pwy similarity'] = {"size": 25 * len(df_clean) ** -0.25, "style": 'italic'}
    annot_kws['pheno similarity'] = {"size": 25 * len(df_clean) ** -0.25, "style": 'italic'}
    annot_kws['dig score'] = {"size": 25 * len(df_clean) ** -0.25, "style": 'italic'}
    annot_kws['dig percentile'] = {"size": 25 * len(df_clean) ** -0.25, "style": 'italic'}
    annot_kws['coex rank'] = {"size": 25 * len(df_clean) ** -0.25, "style": 'italic'}

    hmap = seaborn.heatmap(df_clean, ax=ax, vmin=vmins[feat], vmax=vmaxs[feat],
                           linewidths=4, linecolor=background_color, cbar=False,
                           square=True, cmap=cmaps[feat], mask=mask, annot=True, fmt=annots[feat],
                           annot_kws=annot_kws[feat],
                           cbar_kws={"shrink": .4, 'ticks': ticks[feat], 'spacing': 'proportional'})

    ax.set_xticks([i + 0.5 for i in range(len(df_clean.index) - 1)])
    ax.set_xticklabels(df_clean.index[:-1], fontsize=12, rotation=70, ha='center')
    ax.set_yticks([i + 1.5 for i in range(len(df_clean.index) - 1)])
    ax.set_yticklabels(df_clean.index[1:], fontsize=12, rotation=30)
    ax.set_xlabel(ax_labels[feat], fontsize=16)
    # ax.set_xlim([0, len(df_clean.index)])
    ax.set_ylim([len(df_clean.index), 1])

    for x, g1 in enumerate(df_clean.index[:-1]):
        for y, g2 in enumerate(df_clean.index[1:]):

            if g1 != g2 and mask[x][y] == True:

                ga, gb = tuple(sorted([g1, g2]))
                inh = str(dig_df.loc[(dig_df['gene A'] == ga) & (dig_df['gene B'] == gb)]['pair inheritance'])

                if 'bi-parental' in inh and 'de novo' in inh:

                    ax.axhspan(ymin=y + 1 + 0.04,
                               ymax=y + 2 - 0.04,
                               xmin=(x + 0.04) / len(df_clean),
                               xmax=(x + 1 - 0.04) / len(df_clean), linewidth=1,
                               edgecolor=bi_parental_de_novo_color, facecolor=None, fill=False,
                               alpha=0.95)

                elif 'bi-parental' in inh:
                    ax.axhspan(ymin=y + 1 + 0.04,
                               ymax=y + 2 - 0.04,
                               xmin=(x + 0.04) / len(df_clean),
                               xmax=(x + 1 - 0.04) / len(df_clean), linewidth=1,
                               edgecolor=bi_parental_color, facecolor=None, fill=False,
                               alpha=0.95)
                elif 'de novo' in inh:
                    ax.axhspan(ymin=y + 1 + 0.04,
                               ymax=y + 2 - 0.04,
                               xmin=(x + 0.04) / len(df_clean),
                               xmax=(x + 1 - 0.04) / len(df_clean), linewidth=1,
                               edgecolor=de_novo_color, facecolor=None, fill=False,
                               alpha=0.95)

                elif 'maternal' in inh:
                    ax.axhspan(ymin=y + 1 + 0.04,
                               ymax=y + 2 - 0.04,
                               xmin=(x + 0.04) / len(df_clean),
                               xmax=(x + 1 - 0.04) / len(df_clean), linewidth=1,
                               edgecolor=maternal_color, facecolor=None, fill=False,
                               alpha=0.95)

                elif 'paternal' in inh:
                    ax.axhspan(ymin=y + 1 + 0.04,
                               ymax=y + 2 - 0.04,
                               xmin=(x + 0.04) / len(df_clean),
                               xmax=(x + 1 - 0.04) / len(df_clean), linewidth=1,
                               edgecolor=paternal_color, facecolor=None, fill=False,
                               alpha=0.95)
                else:
                    ax.axhspan(ymin=y + 1 + 0.02,
                               ymax=y + 2 - 0.02,
                               xmin=(x + 0.02) / len(df_clean),
                               xmax=(x + 1 - 0.02) / len(df_clean), linewidth=1,
                               edgecolor=seq_mis_color, facecolor=None, fill=False,
                               alpha=0.95)

    handles = []
    labels = []

    m_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                  edgecolor=maternal_color, facecolor=None, fill=False, linewidth=3, alpha=1)
    p_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                  edgecolor=paternal_color, facecolor=None, fill=False, linewidth=3, alpha=1)

    bp_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                   edgecolor=bi_parental_color, facecolor=None, fill=False, linewidth=3, alpha=1)
    dn_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                   edgecolor=de_novo_color, facecolor=None, fill=False, linewidth=3, alpha=1)
    both_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                     edgecolor=bi_parental_de_novo_color, facecolor=None, fill=False, linewidth=3,
                                     alpha=1)
    miss_box = mpl.patches.Rectangle(xy=(0, 0), width=0, height=0,
                                     edgecolor=seq_mis_color, facecolor=None, fill=False, linewidth=1,
                                     alpha=1)

    handles.append(m_box)
    labels.append('maternal inheritance')
    handles.append(p_box)
    labels.append('paternal inheritance')
    handles.append(bp_box)
    labels.append('biparental inheritance')
    handles.append(dn_box)
    labels.append('de novo variants')
    handles.append(both_box)
    labels.append('biparental inheritance & de novo')
    handles.append(miss_box)
    labels.append('sequencing data missing for one or both parents')

    fig.legend(handles, labels, loc='upper center', fontsize=20, handlelength=0.7, ncol=3,
               facecolor=background_color, edgecolor=background_color, frameon=False, framealpha=0)
    fig.patch.set_facecolor(background_color)

    plt.subplots_adjust(left=0.125, bottom=0.1, right=0.9, top=1, wspace=0.1, hspace=0.1)
    plt.savefig(digenic_score_graph_name + '.pdf', bbox_inches='tight', facecolor=background_color, edgecolor='none')
    plt.savefig(digenic_score_graph_name + '.svg', bbox_inches='tight', facecolor=background_color, edgecolor='none')

    print('digenic score heatmap filename: {}'.format(digenic_score_graph_name))
    return dig_df, dfs


def make_html(df): ## generate DiGePred scores and gene pair details interactive heatmap and detailed JSON
    dig_info_dict = dict()
    genes = sorted(list(set(df['gene A']).union(df['gene B'])))

    dfs = {'DiGePred': {}, 'systems biology': {}}

    dfs['DiGePred']['score'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes), dtype=float)
    # dfs['DiGePred']['pcile'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes), dtype=float)
    dfs['systems biology']['dir. int.'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes),
                                                       dtype=float)
    dfs['systems biology']['pwy sim.'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes),
                                                      dtype=float)
    dfs['systems biology']['pheno sim.'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes),
                                                        dtype=float)
    dfs['systems biology']['coex.'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes),
                                                   dtype=float)

    data_ranges = {'DiGePred': {}, 'systems biology': {}}
    color_scales = {'DiGePred': {}, 'systems biology': {}}

    data_ranges['DiGePred']['score'] = [0.01, 0.1, 0.5]
    # data_ranges['DiGePred']['pcile'] = [0.01, 50, 99]
    data_ranges['systems biology']['dir. int.'] = [1]
    data_ranges['systems biology']['pwy sim.'] = [0.01, 10, 50]
    data_ranges['systems biology']['pheno sim.'] = [0.01, 10, 50]
    data_ranges['systems biology']['coex.'] = [2000, 200, 20]

    trace_names = {'DiGePred': {}, 'systems biology': {}}
    trace_names['DiGePred']['score'] = 'DiGePred score'
    # trace_names['DiGePred']['pcile'] = 'percentile of DiGePred score'
    trace_names['systems biology']['dir. int.'] = 'direct interactions'
    trace_names['systems biology']['pwy sim.'] = 'KEGG and Reactome pathway similarity'
    trace_names['systems biology']['pheno sim.'] = 'HPO phenotype similarity'
    trace_names['systems biology']['coex.'] = 'mutual co-expression rank'

    inh_df = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes), dtype=str)

    for i in df.index:
        ga = df['gene A'][i]
        gb = df['gene B'][i]
        score = df['digenic score'][i]

        dfs['DiGePred']['score'][gb][ga] = score
        dfs['DiGePred']['score'][ga][gb] = score

        # dfs['DiGePred']['pcile'][ga][gb] = df['percentile'][i]
        # dfs['DiGePred']['pcile'][gb][ga] = df['percentile'][i]

        dfs['systems biology']['dir. int.'][ga][gb] = [1 if df['PPI distance'][i] == 1 or
                                                            df['pathway distance'][i] == 1 or
                                                            df['literature distance'][i] == 1 else np.nan][0]
        dfs['systems biology']['dir. int.'][gb][ga] = [1 if df['PPI distance'][i] == 1 or
                                                            df['pathway distance'][i] == 1 or
                                                            df['literature distance'][i] == 1 else np.nan][0]

        dfs['systems biology']['pwy sim.'][ga][gb] = df['pathway similarity'][i]
        dfs['systems biology']['pwy sim.'][gb][ga] = df['pathway similarity'][i]

        dfs['systems biology']['pheno sim.'][ga][gb] = df['phenotype similarity'][i]
        dfs['systems biology']['pheno sim.'][gb][ga] = df['phenotype similarity'][i]

        dfs['systems biology']['coex.'][ga][gb] = [df['mutual co-expression rank'][i] if
                                                   df['mutual co-expression rank'][i] > 0 else np.nan][0]
        dfs['systems biology']['coex.'][gb][ga] = [df['mutual co-expression rank'][i] if
                                                   df['mutual co-expression rank'][i] > 0 else np.nan][0]
        inh_df[ga][gb] = df['pair inheritance'][i]
        inh_df[gb][ga] = df['pair inheritance'][i]

    for g in genes:
        dfs['DiGePred']['score'][g][g] = np.nan
        # dfs['DiGePred']['pcile'][g][g] = np.nan
        dfs['systems biology']['dir. int.'][g][g] = np.nan
        dfs['systems biology']['pwy sim.'][g][g] = np.nan
        dfs['systems biology']['pheno sim.'][g][g] = np.nan
        dfs['systems biology']['coex.'][g][g] = np.nan

    best_vals = {'DiGePred': {}, 'systems biology': {}}
    best_vals['DiGePred']['score'] = 0.5
    # best_vals['DiGePred']['pcile'] = 95
    best_vals['systems biology']['dir. int.'] = 1
    best_vals['systems biology']['pwy sim.'] = 10
    best_vals['systems biology']['pheno sim.'] = 10
    best_vals['systems biology']['coex.'] = 500

    color_scales = {'DiGePred': {}, 'systems biology': {}}

    fig = go.Figure()
    datas = {}
    set_details = []
    values = []
    cutoffs = []
    i_fig = 0
    for k in dfs:
        datas[k] = {}
        for s in dfs[k]:
            datas[k][s] = {}

    for k in dfs:
        for s in dfs[k]:
            for i_s, step in enumerate(data_ranges[k][s]):

                step_ratio = (i_s + 1) / float(len(data_ranges[k][s]))

                color_scales['DiGePred']['score'] = [[0.0, 'white'], [1, '#3f007d']]
                # color_scales['DiGePred']['pcile'] = [[0.0, 'white'], [1, '#3f007d']]
                color_scales['systems biology']['dir. int.'] = [[0.0, 'white'], [1.0, '#3f007d']]
                color_scales['systems biology']['pwy sim.'] = [[0.0, 'white'], [1, '#3f007d']]
                color_scales['systems biology']['pheno sim.'] = [[0.0, 'white'], [1, '#3f007d']]
                color_scales['systems biology']['coex.'] = [[0.0, '#3f007d'], [1.0, 'white']]

                wide_df_i = pd.DataFrame(index=[i for i in dfs[k][s].index],
                                         data=dfs[k][s].values,
                                         columns=[c for c in dfs[k][s].columns])
                # if len(wide_df_i) < 30:
                z = wide_df_i.values
                plot_df = wide_df_i.copy()
                vdf = pd.DataFrame(index=plot_df.index,
                                   columns=plot_df.columns,
                                   data=[[str(np.round(wide_df_i[c][r], 3))
                                          for c in plot_df.columns]
                                         for r in plot_df.index])
                vdf = vdf.replace('nan', 'NA')

                if s == 'dir. int.':
                    for row in plot_df.index:
                        for col in plot_df.columns:
                            v = plot_df[col][row]
                            if row != col:
                                ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                                        gb=tuple(sorted([row, col]))[1])

                                if ip not in dig_info_dict:
                                    dig_info_dict[ip] = {'interactions': {'types': [],
                                                                          'sources': [],
                                                                          'papers': []},
                                                         'pathways': {'similarity': 0, 'common': []},
                                                         'phenotypes': {'similarity': 0, 'common': []},
                                                         'coexpression': {'rank': 0, 'tissues': []},
                                                         'DiGePred': {'score': 0, 'percentile': 0}}
                            if v == 1 and row != col:
                                papers = []
                                itypes = []
                                ievidences = []
                                sources = []
                                if (row, col) in pair_to_int_code_dict:
                                    int_codes = pair_to_int_code_dict[(row, col)]
                                elif (col, row) in pair_to_int_code_dict:
                                    int_codes = pair_to_int_code_dict[(col, row)]
                                else:
                                    continue
                                for c in int_codes:
                                    ct = re.search('[a-zA-Z ]+', c)
                                    sources.append(c[ct.start(): ct.end()])
                                    if c in interaction_details_dict:
                                        int_details = interaction_details_dict[c]
                                        i_type = int_details['interaction type'].replace('_', ' ')
                                        i_evidence = int_details['evidence'].replace('_', ' ').replace(
                                            'Cause', '').replace('-Cause', '').replace('-', '').split(':')
                                        if i_type != '':
                                            itypes.append(i_type)
                                        if i_evidence:
                                            ievidences.extend(i_evidence)
                                        pub_ids = int_details['pub ID'].split('|')
                                        for p in pub_ids:
                                            if p in pub_id_details_dict:
                                                pub_details = pub_id_details_dict[p]
                                                f_author = \
                                                    pub_details['authors'].split(';')[0].split(',')[0].split(' ')[0]
                                                year = pub_details['year']
                                                journal = pub_details['journal']
                                                paper = '{fa} et al., {y} ({j})'.format(fa=f_author, j=journal, y=year)
                                                papers.append(paper)

                                itypes = sorted(set(itypes))
                                ievidences = sorted(set(ievidences))
                                if '' in ievidences:
                                    ievidences.remove('')
                                sources = sorted(set(sources))
                                papers = sorted(set(papers))
                                nk = []
                                for pp in papers:
                                    n = re.split(r'(\d+)', pp)[1]
                                    nk.append(n)
                                papers = [p for p, y in sorted(zip(papers, nk),
                                                               key=lambda pair: pair[1], reverse=True)]

                                disp_text = ''

                                if itypes:
                                    disp_text += '<br>' + ';'.join(itypes)
                                    dig_info_dict[ip]['interactions']['types'].extend(itypes)

                                if ievidences:
                                    dig_info_dict[ip]['interactions']['types'].extend(ievidences)

                                    disp_text += '<br>' + '<br>'.join(ievidences)

                                if sources:
                                    dig_info_dict[ip]['interactions']['sources'].extend(sources)

                                    disp_text += '<br>' + ';'.join(sources)

                                if papers:
                                    dig_info_dict[ip]['interactions']['papers'] = papers

                                    if len(papers) > 5:
                                        disp_text += '<br>' + '<br>'.join(papers[:5])
                                    else:
                                        disp_text += '<br>' + '<br>'.join(papers)

                                vdf[row][col] = disp_text
                                vdf[col][row] = disp_text

                                ds = np.round(dfs['DiGePred']['score'][row][col], 3)
                                ds_text = '<br><br>DiGePred score: {ds}'.format(ds=ds)
                                vdf[col][row] += ds_text
                                vdf[row][col] += ds_text

                if 'pwy' in s:
                    for row in plot_df.index:
                        for col in plot_df.columns:
                            if row != col:
                                ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                                        gb=tuple(sorted([row, col]))[1])

                                if ip not in dig_info_dict:
                                    dig_info_dict[ip] = {'interactions': {'types': [],
                                                                          'sources': [],
                                                                          'papers': []},
                                                         'pathways': {'similarity': 0, 'common': []},
                                                         'phenotypes': {'similarity': 0, 'common': []},
                                                         'coexpression': {'rank': 0, 'tissues': []},
                                                         'DiGePred': {'score': 0, 'percentile': 0}}

                            if row != col and plot_df[col][row] != 'NA':
                                disp_text = ''

                                v = np.round(plot_df[col][row], 2)

                                if row in kegg_gene_to_codes_dict:
                                    row_pwys = kegg_gene_to_codes_dict[row]
                                else:
                                    row_pwys = []
                                if row in reactome_gene_to_codes_dict:
                                    row_pwys.extend(reactome_gene_to_codes_dict[row])

                                if col in kegg_gene_to_codes_dict:
                                    col_pwys = kegg_gene_to_codes_dict[col]
                                else:
                                    col_pwys = []
                                if col in reactome_gene_to_codes_dict:
                                    col_pwys.extend(reactome_gene_to_codes_dict[col])

                                if row_pwys and col_pwys:
                                    disp_text += '{v} %'.format(v=v)

                                    com_pwys = list(set(row_pwys).intersection(col_pwys))

                                    com_pwy_names = sorted(
                                        set([kegg_code_to_name_dict[c] for c in com_pwys if
                                             c in kegg_code_to_name_dict] +
                                            [reactome_code_to_name_dict[c] for c in com_pwys if
                                             c in reactome_code_to_name_dict]))

                                    disp_text += '<br>' + '<br>'.join(com_pwy_names)

                                    dig_info_dict[ip]['pathways']['similarity'] = v
                                    dig_info_dict[ip]['pathways']['common'] = com_pwy_names

                                else:
                                    disp_text += 'NA'

                                vdf[row][col] = disp_text
                                vdf[col][row] = disp_text

                                ds = np.round(dfs['DiGePred']['score'][row][col], 3)
                                ds_text = '<br><br>DiGePred score: {ds}'.format(ds=ds)
                                vdf[col][row] += ds_text
                                vdf[row][col] += ds_text

                if 'pheno' in s:
                    for row in plot_df.index:
                        for col in plot_df.columns:
                            if row != col:
                                ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                                        gb=tuple(sorted([row, col]))[1])

                                if ip not in dig_info_dict:
                                    dig_info_dict[ip] = {'interactions': {'types': [],
                                                                          'sources': [],
                                                                          'papers': []},
                                                         'pathways': {'similarity': 0, 'common': []},
                                                         'phenotypes': {'similarity': 0, 'common': []},
                                                         'coexpression': {'rank': 0, 'tissues': []},
                                                         'DiGePred': {'score': 0, 'percentile': 0}}

                            if row != col and plot_df[col][row] != 'NA':
                                disp_text = ''

                                v = np.round(plot_df[col][row], 2)

                                if row in pheno_gene_to_codes_dict:
                                    row_phenos = pheno_gene_to_codes_dict[row]
                                else:
                                    row_phenos = []
                                if col in pheno_gene_to_codes_dict:
                                    col_phenos = pheno_gene_to_codes_dict[col]
                                else:
                                    col_phenos = []

                                if row_phenos and col_phenos:
                                    disp_text += '{v} %'.format(v=v)

                                    com_phenos = list(set(row_phenos).intersection(col_phenos))

                                    dig_info_dict[ip]['phenotypes']['similarity'] = v
                                    dig_info_dict[ip]['phenotypes']['common'] = \
                                        sorted(set([pheno_code_to_name_dict[c]
                                                    for c in set(com_phenos) if
                                                    c in pheno_code_to_name_dict and
                                                    'obsolete' not in pheno_code_to_name_dict[
                                                        c].lower()]))

                                    if len(com_phenos) > 10:
                                        nx = []
                                        for p in com_phenos:
                                            if p in pheno_tree:
                                                nx.append(list(pheno_tree[p])[
                                                              [-4 if len(pheno_tree[p]) > 4 else 0][0]])

                                        com_phenos = nx

                                    com_pheno_names = sorted(set([pheno_code_to_name_dict[c]
                                                                  for c in com_phenos if
                                                                  c in pheno_code_to_name_dict and
                                                                  'obsolete' not in pheno_code_to_name_dict[
                                                                      c].lower()]))

                                    disp_text += '<br>' + '<br>'.join(com_pheno_names)

                                else:
                                    disp_text += 'NA'

                                vdf[row][col] = disp_text
                                vdf[col][row] = disp_text

                                ds = np.round(dfs['DiGePred']['score'][row][col], 3)
                                ds_text = '<br><br>DiGePred score: {ds}'.format(ds=ds)
                                vdf[col][row] += ds_text
                                vdf[row][col] += ds_text

                if k == 'DiGePred':
                    for row in plot_df.index:
                        for col in plot_df.columns:
                            if row != col:
                                ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                                        gb=tuple(sorted([row, col]))[1])

                                if ip not in dig_info_dict:
                                    dig_info_dict[ip] = {'interactions': {'types': [],
                                                                          'sources': [],
                                                                          'papers': []},
                                                         'pathways': {'similarity': 0, 'common': []},
                                                         'phenotypes': {'similarity': 0, 'common': []},
                                                         'coexpression': {'rank': 0, 'tissues': []},
                                                         'DiGePred': {'score': 0, 'percentile': 0}}

                            if row != col and plot_df[row][col] != 'NA':

                                v = np.round(plot_df[row][col], 2)
                                gp = tuple(sorted([row, col]))
                                dig_data = df.loc[(df['gene A'] == gp[0]) & (df['gene B'] == gp[1])].squeeze()

                                disp_text = ''

                                if s == 'score':
                                    dig_info_dict[ip]['DiGePred']['score'] = v
                                else:
                                    dig_info_dict[ip]['DiGePred']['percentile'] = v

                                if len(dig_data) > 0:
                                    pwy_sim = int(round(dig_data['pathway similarity'], 0))
                                    pheno_sim = int(round(dig_data['phenotype similarity'], 0))

                                    dists = []
                                    if dig_data['PPI distance'] and dig_data['PPI distance'] > 0:
                                        dists.append(dig_data['PPI distance'])
                                    if dig_data['pathway distance'] and dig_data['pathway distance'] > 0:
                                        dists.append(dig_data['pathway distance'])
                                    if dig_data['literature distance'] and dig_data['literature distance'] > 0:
                                        dists.append(dig_data['literature distance'])

                                    if dists:
                                        net_dist = int(np.min(dists))
                                    else:
                                        net_dist = 0

                                    coex = int(round(dig_data['mutual co-expression rank'], 0))

                                    if 'pair inheritance' in df.columns:
                                        inh = df.loc[(df['gene A'] == gp[0]) &
                                                     (df['gene B'] == gp[1])].squeeze()['pair inheritance']
                                        if inh != "['NA']":
                                            if inh == 'both parents not sequenced':
                                                disp_text += '<br>{inh}'.format(inh=inh)
                                            elif 'de novo' not in inh:
                                                disp_text += '<br>{inh} inheritance'.format(inh=inh)
                                            else:
                                                disp_text += '<br><i>de novo</i>'

                                    if pwy_sim > 0:
                                        disp_text += '<br>pathway similarity: {pw}% '.format(pw=pwy_sim)
                                    if pheno_sim > 0:
                                        disp_text += '<br>phenotype similarity: {pw}% '.format(pw=pheno_sim)
                                    if net_dist == 1:
                                        disp_text += '<br>direct interaction'
                                    elif net_dist > 0:
                                        disp_text += '<br>{pw} proteins away'.format(pw=net_dist)
                                    if coex > 0:
                                        disp_text += '<br>mutual co-expression rank: {pw}'.format(pw=coex)

                                vdf[row][col] = '{v}'.format(v=v) + disp_text
                                vdf[col][row] = '{v}'.format(v=v) + disp_text

                                ds = np.round(dfs['DiGePred']['score'][row][col], 3)
                                ds_text = '<br><br>DiGePred score: {ds}'.format(ds=ds)
                                vdf[col][row] += ds_text
                                vdf[row][col] += ds_text

                if s == 'coex.':
                    for row in plot_df.index:
                        for col in plot_df.columns:
                            if row != col:
                                ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                                        gb=tuple(sorted([row, col]))[1])

                                if ip not in dig_info_dict:
                                    dig_info_dict[ip] = {'interactions': {'types': [],
                                                                          'sources': [],
                                                                          'papers': []},
                                                         'pathways': {'similarity': 0, 'common': []},
                                                         'phenotypes': {'similarity': 0, 'common': []},
                                                         'coexpression': {'rank': 0, 'tissues': []},
                                                         'DiGePred': {'score': 0, 'percentile': 0}}

                            if row != col and plot_df[row][col] != 'NA' and plot_df[row][col] > 0:
                                v = int(np.round(plot_df[row][col], 0))

                                disp_text = ''

                                if row in gene_to_mean_TPM_by_tissue_dict:
                                    tpmsa = np.asarray(gene_to_mean_TPM_by_tissue_dict[row])
                                    if np.sum(tpmsa) > 0:
                                        tpmsa = tpmsa / np.sum(tpmsa)
                                        if col in gene_to_mean_TPM_by_tissue_dict:
                                            tpmsb = np.asarray(gene_to_mean_TPM_by_tissue_dict[col])
                                            if np.sum(tpmsb) > 0:
                                                tpmsb = tpmsb / np.sum(tpmsb)

                                        highest_ex_tissues_a = []
                                        highest_ex_tissues_b = []
                                        tpm_vals_dict = dict()

                                        for i, vv in sorted([(i, x) for (i, x) in enumerate(tpmsa)],
                                                            key=lambda x: x[1], reverse=True):
                                            if vv > 0:
                                                highest_ex_tissues_a.append(id_to_narrow_tissue_dict[i])

                                            tpm_vals_dict[id_to_narrow_tissue_dict[i]] = vv

                                        for i, vv in sorted([(i, x) for (i, x) in enumerate(tpmsb)],
                                                            key=lambda x: x[1], reverse=True):
                                            if vv > 0:
                                                highest_ex_tissues_b.append(id_to_narrow_tissue_dict[i])

                                            tpm_vals_dict[id_to_narrow_tissue_dict[i]] += vv

                                        highest_ex_tissues_com = set(highest_ex_tissues_a[:10]).intersection(
                                            highest_ex_tissues_b[:10])
                                        highest_ex_tissues_com = sorted(highest_ex_tissues_com,
                                                                        key=lambda x: tpm_vals_dict[x],
                                                                        reverse=True)

                                        if highest_ex_tissues_com:
                                            disp_text += '<br>' + '<br>'.join(highest_ex_tissues_com)

                                vdf[row][col] = '{v}'.format(v=v) + disp_text
                                vdf[col][row] = '{v}'.format(v=v) + disp_text

                                dig_info_dict[ip]['coexpression']['rank'] = v
                                dig_info_dict[ip]['coexpression']['tissues'] = highest_ex_tissues_com

                                ds = np.round(dfs['DiGePred']['score'][row][col], 3)
                                ds_text = '<br><br>DiGePred score: {ds}'.format(ds=ds)
                                vdf[col][row] += ds_text
                                vdf[row][col] += ds_text

                if s != 'coex.':
                    plot_df[plot_df < step] = 0
                    plot_df = plot_df.fillna(value=0)
                else:
                    plot_df[plot_df > step] = 2000
                    plot_df = plot_df.fillna(value=2000)
                z2 = plot_df.values
                if len(z2) == 0 or np.max(np.max(plot_df)) == 0 or np.min(np.min(plot_df)) == 2000:
                    fig.add_trace(
                        go.Heatmap(visible=False, name=trace_names[k][s], opacity=0.9,
                                   x=[c for c in plot_df.columns],
                                   y=[i for i in plot_df.index],
                                   z=z2, text=vdf,
                                   type='heatmap', hoverinfo='text+x+y', hovertemplate="<b>%{text}</b><br>" +
                                                                                       "gene A: %{x}<br>" +
                                                                                       "gene B: %{y}<br>"
                                                                                       "<extra></extra>"
                                   ,
                                   colorscale=[[0.0, 'white'], [1.0, 'white']], xperiodalignment='start',
                                   colorbar=dict(tickmode='array', tickvals=data_ranges[k][s]),
                                   )
                    )
                    datas[k][s][step] = i_fig
                    set_details.append((k, s, step))
                    i_fig += 1
                else:

                    datas[k][s][step] = z2
                    cutoffs.append(step)
                    values.append(plot_df)
                    fig.add_trace(
                        go.Heatmap(visible=False, name=trace_names[k][s], opacity=0.9,
                                   x=[c for c in plot_df.columns],
                                   y=[i for i in plot_df.index],
                                   z=z2, text=vdf,
                                   type='heatmap', hoverinfo='text+x+y', hovertemplate="<b>%{text}</b><br>" +
                                                                                       "gene A: %{x}<br>" +
                                                                                       "gene B: %{y}<br>"
                                                                                       "<extra></extra>"
                                   ,
                                   colorscale=color_scales[k][s], xperiodalignment='start',
                                   colorbar=dict(tickmode='array', tickvals=data_ranges[k][s]),
                                   )
                    )
                    datas[k][s][step] = i_fig
                    set_details.append((k, s, step))
                    i_fig += 1

    fig.data[1].visible = True
    fig.update_traces(showscale=False)

    net_details = ['']

    vis_data = {}
    vis_dicts = {}
    all_vis_dicts = []
    for i, k in enumerate(dfs):
        vis_data[k] = {}
        vis_dicts[k] = {}
        for j, s in enumerate(dfs[k]):
            vis_data[k][s] = {}
            vis_dicts[k][s] = []
            nn = 0
            for st in datas[k][s]:
                vis_data[k][s][st] = [False] * len(fig.data)
                n = datas[k][s][st]
                vis_data[k][s][st][n] = True
                if 'pwy' in s or 'pheno' in s:
                    all_vis_dicts.append(dict(args=[{'visible': vis_data[k][s][st]},
                                                    {'title': '{n} | cutoff >= {c}'.format(n=trace_names[k][s],
                                                                                           c=[st if st > 0.01 else 0][
                                                                                               0]),
                                                     'font_size': 16,
                                                     'font_family': "Arial"}],
                                              args2=[{'visible': [False] * len(fig.data)},
                                                     {'title': 'Please select a metric',
                                                      'font_size': 16,
                                                      'font_family': "Arial"}],
                                              label=['\u2265 {v} %'.format(v=st) if st > 0.01 else 'ALL'][0],
                                              method='update',
                                              ))
                elif s != 'dir. int.' and s != 'coex.':
                    all_vis_dicts.append(dict(args=[{'visible': vis_data[k][s][st]},
                                                    {'title': '{n} | cutoff >= {c}'.format(n=trace_names[k][s],
                                                                                           c=[st if st > 0.01 else 0][
                                                                                               0]),
                                                     'font_size': 16,
                                                     'font_family': "Arial"}],
                                              args2=[{'visible': [False] * len(fig.data)},
                                                     {'title': 'Please select a metric',
                                                      'font_size': 16,
                                                      'font_family': "Arial"}],
                                              label=['\u2265 {v}'.format(v=st) if st > 0.01 else 'ALL'][0],
                                              method='update',
                                              ))
                elif s == 'dir. int.':
                    all_vis_dicts.append(dict(args=[{'visible': vis_data[k][s][st]},
                                                    {'title': 'direct interactions (PPI, pathway & literature-mined interactions)',
                                                     'font_size': 16,
                                                     'font_family': "Arial"}
                                                    ],
                                              label='direct interactions',
                                              method='update',
                                              args2=[{'visible': [False] * len(fig.data)},
                                                     {'title': 'Please select a metric',
                                                      'font_size': 16,
                                                      'font_family': "Arial"}]
                                              ))
                else:
                    all_vis_dicts.append(dict(args=[{'visible': vis_data[k][s][st]},
                                                    {'title': '{n} | cutoff <= {c}'.format(n=trace_names[k][s], c=st),
                                                     'font_size': 16,
                                                     'font_family': "Arial"},
                                                    ],
                                              args2=[{'visible': [False] * len(fig.data)},
                                                     {'title': 'Please select a metric',
                                                      'font_size': 16,
                                                      'font_family': "Arial"}],
                                              label='\u2264 {v}'.format(v=st),
                                              method='update',
                                              ))

    inh_colors = {'de novo': '#fdb863',
                  'bi-parental': '#5aae61',
                  'maternal': '#fff7bc',
                  'paternal': '#9ecae1',
                  'NA': '#969696',
                  }

    for i, ga in enumerate(inh_df.index):
        for j, gb in enumerate(inh_df.columns):
            if ga != gb:
                inh = str(inh_df[ga][gb])
                if 'de novo' in inh:
                    clr = inh_colors['de novo']
                    fig.add_shape(type="rect",
                                  x0=j - 0.45, y0=i - 0.45, x1=j + 0.45, y1=i + 0.45,
                                  line=dict(
                                      color=clr,
                                      width=2,
                                  ),
                                  )
                elif 'bi-parental' in inh:
                    clr = inh_colors['bi-parental']
                    fig.add_shape(type="rect",
                                  x0=j - 0.45, y0=i - 0.45, x1=j + 0.45, y1=i + 0.45,
                                  line=dict(
                                      color=clr,
                                      width=2,
                                  ),
                                  )
                elif 'paternal' in inh:
                    clr = inh_colors['paternal']
                    fig.add_shape(type="rect",
                                  x0=j - 0.45, y0=i - 0.45, x1=j + 0.45, y1=i + 0.45,
                                  line=dict(
                                      color=clr,
                                      width=1,
                                  ),
                                  )
                elif 'maternal' in inh:
                    clr = inh_colors['maternal']
                    fig.add_shape(type="rect",
                                  x0=j - 0.45, y0=i - 0.45, x1=j + 0.45, y1=i + 0.45,
                                  line=dict(
                                      color=clr,
                                      width=1,
                                  ),
                                  )
                else:
                    clr = inh_colors['NA']
                    fig.add_shape(type="rect",
                                  x0=j - 0.45, y0=i - 0.45, x1=j + 0.45, y1=i + 0.45,
                                  line=dict(
                                      color=clr,
                                      width=0.5,
                                  ),
                                  )
            else:
                clr = inh_colors['NA']
                fig.add_shape(type="rect",
                              x0=j - 0.45, y0=i - 0.45, x1=j + 0.45, y1=i + 0.45,
                              line=dict(
                                  color=clr,
                                  width=0.5,
                              ),
                              )

    fig.update_layout(
        updatemenus=[
            dict(showactive=True, bgcolor='#bdbdbd', bordercolor='#636363',
                 type="buttons", font_size=10, font_family="Arial",
                 direction="right",
                 active=1,
                 pad={"t": 10}, x=0.89, y=-0.17,
                 buttons=list(all_vis_dicts),
                 ),

            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['maternal'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.9,
                 buttons=[dict(args=[{}], label='maternal only', method='skip')]),
            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['paternal'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.8,
                 buttons=[dict(args=[{}], label='paternal only', method='skip')]),

            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['bi-parental'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.7,
                 buttons=[dict(args=[{}], label='bi-parental', method='skip')]),
            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['de novo'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.6,
                 buttons=[dict(args=[{}], label='<i>de novo</i>', method='skip')]),

            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['NA'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.5,
                 buttons=[dict(args=[{}], label='NA', method='skip')]),

        ])
    fig.update_layout(
        width=1000, height=750,
        autosize=False,
        margin=dict(t=100, b=200, l=100, r=100, pad=0),
    )

    fig.update_layout(
        annotations=[
            dict(text="DiGePred<br>score", x=0.039, xref="paper", y=-0.355, yref="paper", font_size=14,
                 font_family="Arial",
                 align="center", showarrow=False),
            dict(text="   PPI   <br>     ", x=0.21, xref="paper", y=-0.355, yref="paper", font_size=14,
                 font_family="Arial",
                 align="center", showarrow=False),
            dict(text="KEGG and Reactome<br>pathway similarity", x=0.4, xref="paper", y=-0.355, yref="paper",
                 font_size=14,
                 font_family="Arial",
                 align="center", showarrow=False),
            dict(text="HPO phenotype<br>similarity", x=0.6, xref="paper", y=-0.355, yref="paper", font_size=14,
                 font_family="Arial",
                 align="center", showarrow=False),
            dict(text="co-expression<br>rank", x=0.85, xref="paper", y=-0.355, yref="paper", font_size=14,
                 font_family="Arial",
                 align="center", showarrow=False),
            dict(text="Select Data", x=1.02, xref="paper", y=-0.26, yref="paper", font_size=18,
                 font_family="Arial",
                 align="left", showarrow=False),
            dict(text="|", x=0.165, xref="paper", y=-0.33, yref="paper", font_size=18,
                 font_family="Arial",
                 align="left", showarrow=False),
            dict(text="|", x=0.295, xref="paper", y=-0.33, yref="paper", font_size=18,
                 font_family="Arial",
                 align="left", showarrow=False),
            dict(text="|", x=0.5, xref="paper", y=-0.33, yref="paper", font_size=18,
                 font_family="Arial",
                 align="left", showarrow=False),
            dict(text="|", x=0.71, xref="paper", y=-0.33, yref="paper", font_size=18,
                 font_family="Arial",
                 align="left", showarrow=False),
            dict(text="Inheritance", x=0.89, xref="paper", y=0.95, yref="paper", font_size=18,
                 font_family="Arial",
                 align="left", showarrow=False),
        ]
    )

    fig.update_layout(xaxis_showgrid=True, yaxis_showgrid=True,
                      title_text="DiGePred score | cutoff >= 0.1", hoverlabel=dict(
            bgcolor="white",
            font_size=16,
            font_family="Arial"
        ),
                      )
    fig.update_xaxes(tickson='boundaries', showgrid=True, gridcolor='black', gridwidth=2, zeroline=True,
                     showspikes=True, spikecolor='#74c476', spikethickness=2, spikedash='dot',
                     spikemode='toaxis+marker')
    fig.update_yaxes(tickson='boundaries', showgrid=True, gridcolor='black', gridwidth=2, zeroline=True,
                     showspikes=True, spikecolor='#74c476', spikethickness=2, spikedash='dot',
                     spikemode='toaxis+marker')

    fig['layout']['xaxis']['scaleanchor'] = 'y'
    fig['layout']['xaxis']['constrain'] = 'domain'
    fig['layout']['xaxis']['constraintoward'] = 'left'
    fig['layout']['xaxis']['scaleratio'] = 1

    pio.write_html(fig, file=html_file_name)
    print('html filename: {}'.format(os.path.abspath(html_file_name)))
    json.dump(dig_info_dict, open(details_json_name, 'w'), indent=6)
    print('summary json filename: {}'.format(os.path.abspath(details_json_name)))


def make_json(df): ## generate detailed JSON if no. of gene pairs too high to draw heatmap
    dig_info_dict = dict()
    genes = sorted(list(set(df['gene A']).union(df['gene B'])))

    for i in df.index:
        ga = df['gene A'][i]
        gb = df['gene B'][i]
        score = '%0.3g' % float(df['digenic score'][i])
        pcile = df['percentile'][i]

        dir_int = [1 if df['PPI distance'][i] == 1 or
                        df['pathway distance'][i] == 1 or
                        df['literature distance'][i] == 1 else np.nan][0]
        coex = np.round([df['mutual co-expression rank'][i] if
                         df['mutual co-expression rank'][i] > 0 else np.nan][0], 0)

        inh = df['pair inheritance'][i]
        pwy = np.round(df['pathway similarity'][i], 2)
        pheno = np.round(df['phenotype similarity'][i], 2)

        row = ga
        col = gb

        if row != col:

            ip = '{ga},{gb}'.format(ga=ga, gb=gb)

            if ip not in dig_info_dict:
                dig_info_dict[ip] = {'interactions': {'types': [],
                                                      'sources': [],
                                                      'papers': []},
                                     'pathways': {'similarity': 0, 'common': []},
                                     'phenotypes': {'similarity': 0, 'common': []},
                                     'coexpression': {'rank': 0, 'tissues': []},
                                     'DiGePred': {'score': 0, 'percentile': 0}}
            if dir_int == 1:
                papers = []
                itypes = []
                ievidences = []
                sources = []
                if (row, col) in pair_to_int_code_dict:
                    int_codes = pair_to_int_code_dict[(row, col)]
                elif (col, row) in pair_to_int_code_dict:
                    int_codes = pair_to_int_code_dict[(col, row)]
                else:
                    continue
                for c in int_codes:
                    ct = re.search('[a-zA-Z ]+', c)
                    sources.append(c[ct.start(): ct.end()])
                    if c in interaction_details_dict:
                        int_details = interaction_details_dict[c]
                        i_type = int_details['interaction type'].replace('_', ' ')
                        i_evidence = int_details['evidence'].replace('_', ' ').replace(
                            'Cause', '').replace('-Cause', '').replace('-', '').split(':')
                        if i_type != '':
                            itypes.append(i_type)
                        if i_evidence:
                            ievidences.extend(i_evidence)
                        pub_ids = int_details['pub ID'].split('|')
                        for p in pub_ids:
                            if p in pub_id_details_dict:
                                pub_details = pub_id_details_dict[p]
                                f_author = \
                                    pub_details['authors'].split(';')[0].split(',')[0].split(' ')[0]
                                year = pub_details['year']
                                journal = pub_details['journal']
                                paper = '{fa} et al., {y} ({j})'.format(fa=f_author, j=journal, y=year)
                                papers.append(paper)

                itypes = sorted(set(itypes))
                ievidences = sorted(set(ievidences))
                if '' in ievidences:
                    ievidences.remove('')
                sources = sorted(set(sources))
                papers = sorted(set(papers))
                nk = []
                for pp in papers:
                    n = re.split(r'(\d+)', pp)[1]
                    nk.append(n)
                papers = [p for p, y in sorted(zip(papers, nk),
                                               key=lambda pair: pair[1], reverse=True)]
                if itypes:
                    dig_info_dict[ip]['interactions']['types'].extend(itypes)

                if ievidences:
                    dig_info_dict[ip]['interactions']['types'].extend(ievidences)

                if sources:
                    dig_info_dict[ip]['interactions']['sources'].extend(sources)

                if papers:
                    dig_info_dict[ip]['interactions']['papers'] = papers

                if row in kegg_gene_to_codes_dict:
                    row_pwys = kegg_gene_to_codes_dict[row]
                else:
                    row_pwys = []
                if row in reactome_gene_to_codes_dict:
                    row_pwys.extend(reactome_gene_to_codes_dict[row])

                if col in kegg_gene_to_codes_dict:
                    col_pwys = kegg_gene_to_codes_dict[col]
                else:
                    col_pwys = []
                if col in reactome_gene_to_codes_dict:
                    col_pwys.extend(reactome_gene_to_codes_dict[col])

                if row_pwys and col_pwys:
                    com_pwys = list(set(row_pwys).intersection(col_pwys))

                    com_pwy_names = sorted(
                        set([kegg_code_to_name_dict[c] for c in com_pwys if
                             c in kegg_code_to_name_dict] +
                            [reactome_code_to_name_dict[c] for c in com_pwys if
                             c in reactome_code_to_name_dict]))

                    dig_info_dict[ip]['pathways']['similarity'] = pwy
                    dig_info_dict[ip]['pathways']['common'] = com_pwy_names

                if row in pheno_gene_to_codes_dict:
                    row_phenos = pheno_gene_to_codes_dict[row]
                else:
                    row_phenos = []
                if col in pheno_gene_to_codes_dict:
                    col_phenos = pheno_gene_to_codes_dict[col]
                else:
                    col_phenos = []

                if row_phenos and col_phenos:
                    com_phenos = list(set(row_phenos).intersection(col_phenos))

                    dig_info_dict[ip]['phenotypes']['similarity'] = pheno
                    dig_info_dict[ip]['phenotypes']['common'] = \
                        sorted(set([pheno_code_to_name_dict[c]
                                    for c in set(com_phenos) if
                                    c in pheno_code_to_name_dict and
                                    'obsolete' not in pheno_code_to_name_dict[
                                        c].lower()]))

                dig_info_dict[ip]['DiGePred']['score'] = score
                dig_info_dict[ip]['DiGePred']['percentile'] = pcile

                if row in gene_to_mean_TPM_by_tissue_dict:
                    tpmsa = np.asarray(gene_to_mean_TPM_by_tissue_dict[row])
                    if np.sum(tpmsa) > 0:
                        tpmsa = tpmsa / np.sum(tpmsa)
                        if col in gene_to_mean_TPM_by_tissue_dict:
                            tpmsb = np.asarray(gene_to_mean_TPM_by_tissue_dict[col])
                            if np.sum(tpmsb) > 0:
                                tpmsb = tpmsb / np.sum(tpmsb)

                        highest_ex_tissues_a = []
                        highest_ex_tissues_b = []
                        tpm_vals_dict = dict()

                        for i, vv in sorted([(i, x) for (i, x) in enumerate(tpmsa)],
                                            key=lambda x: x[1], reverse=True):
                            if vv > 0:
                                highest_ex_tissues_a.append(id_to_narrow_tissue_dict[i])

                            tpm_vals_dict[id_to_narrow_tissue_dict[i]] = vv

                        for i, vv in sorted([(i, x) for (i, x) in enumerate(tpmsb)],
                                            key=lambda x: x[1], reverse=True):
                            if vv > 0:
                                highest_ex_tissues_b.append(id_to_narrow_tissue_dict[i])

                            tpm_vals_dict[id_to_narrow_tissue_dict[i]] += vv

                        highest_ex_tissues_com = set(highest_ex_tissues_a[:10]).intersection(
                            highest_ex_tissues_b[:10])
                        highest_ex_tissues_com = sorted(highest_ex_tissues_com,
                                                        key=lambda x: tpm_vals_dict[x],
                                                        reverse=True)
                        if highest_ex_tissues_com:
                            disp_text += '<br>' + '<br>'.join(highest_ex_tissues_com)

                dig_info_dict[ip]['coexpression']['rank'] = v
                dig_info_dict[ip]['coexpression']['tissues'] = highest_ex_tissues_com

    json.dump(dig_info_dict, open(details_json_name, 'w'), indent=6)
    print('summary json filename: {}'.format(os.path.abspath(details_json_name)))


if __name__ == '__main__':

    # get DiGePred df
    dig_df, sel_gp_csv_df = get_digenic_metrics_csv(gene_pairs=pairs)

    # heatmaps and html graphs difficult to read if no. of gene pairs > 500, so only DF and JSON made in that case.
    if len(dig_df.index) < 500:
        plot_heatmaps(dig_df=dig_df, sel_gp_csv_df=sel_gp_csv_df)
        make_html(dig_df)
    else:
        make_json(dig_df)